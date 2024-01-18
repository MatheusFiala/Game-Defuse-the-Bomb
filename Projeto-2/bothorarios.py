import openpyxl

def transferir_horarios(origem, destino):
    # Carregar a planilha de origem
    planilha_origem = openpyxl.load_workbook(origem)
    folha_origem = planilha_origem.active

    # Criar/Carregar a planilha de destino
    try:
        planilha_destino = openpyxl.load_workbook(destino)
    except FileNotFoundError:
        planilha_destino = openpyxl.Workbook()
    folha_destino = planilha_destino.active

    # Iterar sobre as linhas na planilha de origem
    for linha in range(2, folha_origem.max_row + 1):  # Começar da segunda linha assumindo cabeçalho
        # Obter o horário da coluna B (coluna 2)
        horario = folha_origem.cell(row=linha, column=2).value

        # Transferir o horário para a planilha de destino
        folha_destino.append([horario])

    # Salvar a planilha de destino
    planilha_destino.save(destino)

# Exemplo de uso
planilha_origem = '172.32.1.21\desenvolvimento\Registro Ponto/ACOMPANHAMENTO FOLHA PONTO - COLOMBO - Atalho.xlsx'
planilha_destino = '172.32.1.21\desenvolvimento\Registro Ponto2023\DEV/Controle de Horas 2023.xlsx'

transferir_horarios(planilha_origem, planilha_destino)
